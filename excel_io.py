# excel_io.py
# This file handles reading from and writing to Excel files.

import json
import platform
import pandas as pd
import os
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from data_structures import Task, Resource
from utils import resource_path



def read_tasks_from_excel(filename, sheet_name="Sheet1"):
    """Read tasks from an Excel file and return a list of Task objects."""
    df = pd.read_excel(filename)
    tasks = []
    
    for _, row in df.iterrows():
        task_name = row.iloc[0] #  Task name is the first column
        priority = row.iloc[1]  #  priority is the second column
        for col in df.columns[2:]:  # Start from the third column for durations
            if pd.notna(row[col]):
                durations = {col: row[col]}
                full_task_name = f"{col}: {task_name}"
                task = Task(full_task_name, durations, priority)
                tasks.append(task)
    def load_dependency_rules(filename):
        with open(filename, 'r') as file:
            return json.load(file)
    dependency_rules = load_dependency_rules('dependency_rules.json')
    set_up_dependencies(tasks, dependency_rules)
    return tasks


def set_up_dependencies(tasks, rules):
    """Set up dependencies between tasks based on dynamic rules."""
    def get_base_task_name(task):
        return task.name.split(": ")[1].split(" (")[0]

    def find_dependent_tasks(task, keywords):
        base_task_name = get_base_task_name(task)
        return [
            t for t in tasks
            if any(keyword in t.durations for keyword in keywords) and get_base_task_name(t) == base_task_name
        ]

    def add_dependencies(task, keywords):
        dependent_tasks = find_dependent_tasks(task, keywords)
        for dependent in dependent_tasks:
            dependent.dependencies.append(task)

    for task in tasks:
        task_keywords = [key for key, value in rules.items() if key in task.durations]
        for keyword in task_keywords:
            dependent_keywords = rules[keyword]
            add_dependencies(task, dependent_keywords)


            
def read_resources_from_excel(filename, sheet_name="Sheet2"):
    """Read resources from an Excel file, processing all rows in the first column, then the second, and so on."""
    df = pd.read_excel(filename, sheet_name=sheet_name)
    resources = []
    
    for department in df.columns:
        for resource_name in df[department].dropna():
            resource = Resource(resource_name, department)
            resources.append(resource)
    
    return resources

def read_departments_from_excel(filename, sheet_name="Sheet1"):
    df = pd.read_excel(filename, sheet_name=sheet_name)
    return list(df.columns[2:])

def generate_excel_output(schedule, filename):
    """Generates an Excel schedule report with merged department headers and automatically opens it.

    Args:
        schedule: Scheduling data object.
        filename: Name of the output Excel file.
    """
    # Create a list of department names corresponding to each resource
    departments = ['Days'] + [r.department for r in schedule.resources]

    # Create the DataFrame with an extra row for departments
    df = pd.DataFrame(columns=['Days'] + [r.name for r in schedule.resources])

    max_day = max((a[1] for assignments in schedule.assignments.values() for a in assignments), default=0)

    task_colors = {}
    for assignments in schedule.assignments.values():
        for task, _, _ in assignments:
            base_task_name = task.name.split(": ")[1].split(" (")[0]
            if base_task_name not in task_colors:
                task_colors[base_task_name] = ''.join(random.choices('ABCDEF0123456789', k=6))

    for day in range(1, max_day + 1):
        row = {'Days': day}
        for resource in schedule.resources:
            tasks = [f"{a[0].name} ({a[2]:.1f})" for a in schedule.assignments[resource.name] if a[1] == day]
            row[resource.name] = '\n'.join(tasks) if tasks else ''
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    # Write data to Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Schedule', startrow=1)  # Start writing from row 2
    worksheet = writer.sheets['Schedule']

    # Insert the department names row at the top
    current_dept = None
    merge_start = 2  # Start from column 2 since column 1 is "Days"
    for col, dept in enumerate(departments[1:], start=2):  # Skip the "Days" column
        if dept != current_dept:
            if current_dept is not None:
                if merge_start < col - 1:
                    worksheet.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col-1)
                cell = worksheet.cell(row=1, column=merge_start, value=current_dept)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            current_dept = dept
            merge_start = col
        
        if col == len(departments):  # Handle the last department
            if merge_start <= col:
                worksheet.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=col)
            cell = worksheet.cell(row=1, column=merge_start, value=current_dept)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set cell styles for data rows
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(2, len(schedule.resources) + 2):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in range(3, max_day + 3):  # Start from row 3 due to the new department and resource name rows
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                base_task_name = cell.value.split(": ")[1].split(" (")[0]
                if base_task_name in task_colors:
                    cell.fill = PatternFill(start_color=task_colors[base_task_name], end_color=task_colors[base_task_name], fill_type="solid")
                else:
                    print(f"Warning: No color found for task '{base_task_name}'")
            cell.alignment = center_alignment
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(max_length * 1.2, 15)

    max_days_length = max(df['Days'].astype(str).map(len).max(), len('Days')) * 1.2
    worksheet.column_dimensions['A'].width = max_days_length

    # Define border style
    border_style = Border(left=Side(style='thin'), 
                          right=Side(style='thin'), 
                          top=Side(style='thin'), 
                          bottom=Side(style='thin'))

    # Apply border style to each cell, including the new department and resource name rows
    for row in worksheet.iter_rows(min_row=1, max_row=max_day + 2, min_col=1, max_col=len(schedule.resources) + 1):
        for cell in row:
            cell.border = border_style

    # Save the workbook
    writer._save()

    # Use platform-specific commands to open the Excel file automatically
    if platform.system() == "Windows":
        os.startfile(filename)
    elif platform.system() == "Darwin":  # macOS
        os.system("open " + filename)
    else:
        print(f"Warning: Automatic opening not supported for your platform. Please open {filename} manually.")

    print(f"Schedule exported to {filename}")